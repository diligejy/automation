import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
scope = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive'
]
credentials = ServiceAccountCredentials.from_json_keyfile_name('cred.json', scope)
gs = gspread.authorize(credentials)
spread = gs.create('자동화로 생성된 문서')
spread.share('alghost.lee@gmail.com', perm_type='user', role='owner')

#data_path = 'C:\\python\\examples\\2.4\\data'
data_path = './'
file_list = os.listdir(data_path)

for fname in file_list:
    if fname[-4:] != 'xlsx':
        continue
    file_path = os.path.join(data_path, fname)
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    row_count = ws.max_row
    col_count = ws.max_column

    worksheet = spread.add_worksheet(fname, row_count, col_count)
    cells = worksheet.range('A1:' + get_column_letter(col_count) + str(row_count))

    idx = 0
    for row in ws.iter_rows():
        for cell in row:
            cells[idx].value = str(cell.value)
            idx = idx + 1

    worksheet.update_cells(cells)

#spread.share('직원0@gmail.com', perm_type='user', role='reader')
#spread.share('직원1@gmail.com', perm_type='user', role='writer')
#spread.share('직원2@gmail.com', perm_type='user', role='reader')

